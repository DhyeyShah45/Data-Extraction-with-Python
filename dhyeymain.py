"""
The module gives the output to the Feature of Top 3 or Bottom 3 performer
"""
import pandas as pd
from openpyxl import load_workbook


class Employee:
    """
    class Employee
    """
    def __init__(self, e_ps, e_name):
        """
        Function to take inputs
        :param e_ps: PS Number of Employee
        :param e_name: Name of Employee
        """
        self.Emp_Ps = e_ps
        self.Emp_Name = e_name
        self.Emp_marks = {}
        self.Total_Marks = 0


def sheets_access():

    '''
    Returns Sheets Dictionary
    '''
    workbook = load_workbook(filename="Mark_Sheet.xlsx")
    res = workbook.sheetnames
    sheets = {}                                 # empty dictionary
    for i in range(len(res)):                   # for loop for multiple sheets in the excelbook
        sheets[res[i]] = pd.read_excel("Mark_Sheet.xlsx",engine = "openpyxl", sheet_name=res[i])
        sheets[res[i]].dropna(axis=1, how='all', inplace=True)
        sheets[res[i]].dropna(axis=0, how='all', inplace=True)
    return sheets,res

def create_objects(sheets):
    '''
    merge  sheets by marks
    '''

    emp_data = []
    sheets, res = sheets_access()
    for line in range(len(sheets[res[0]]['Marks'])):
        Emp_Name = sheets[res[0]]['Emp Name'][line]
        Emp_Ps = sheets[res[0]]['Emp PS #'][line]
        emp = Employee(Emp_Ps, Emp_Name)
        emp_data.append(emp)

    for i in res:
        count = 0
        for emp in emp_data:
            emp.Emp_marks[i] = float("{:.2f}".format(sheets[i]['Marks'][count]))
            count += 1

    for emp in emp_data:
        emp.Total_Marks = sum(emp.Emp_marks.values())

    return emp_data

def find_top_performer(sheets):
    """
    Finds the top 3 performer
    """
    emp_data = create_objects(sheets)
    new_emp_data = sorted(emp_data, key=lambda x: x.Total_Marks, reverse=True)
    return new_emp_data[0:3] # Top 3 performer

def find_bottom_performer(sheets):
    """
    Finds the bottom 3 performer
    """
    emp_data = create_objects(sheets)
    new_emp_data = sorted(emp_data, key=lambda x: x.Total_Marks, reverse=True)
    return new_emp_data[-4:-1] # Bottom 3 performer

def print_obj(emp):
    """
    Prints the details of the object
    """
    print("Name: " + str(emp.Emp_Name))
    print("PS No. #: " + str(emp.Emp_Ps))
    for key,value in emp.Emp_marks.items():
        print(str(key) + ': ' + str(value))
    print("Total Marks: " + str(emp.Total_Marks))
    print()


def save_data(sorted_list):

    final = pd.DataFrame([l.__dict__ for l in sorted_list ])

    path = r"Mark_Sheet.xlsx"
    book = load_workbook(path)
    writer = pd.ExcelWriter(path, engine='openpyxl')
    writer.book = book
    if 'output' in book.sheetnames:
        pfd = book['output']
        book.remove(pfd)
    final.to_excel(writer, sheet_name='output')
    print("Output written")
    writer.save()
    writer.close()


if __name__=="__main__":

    sheets, res=sheets_access()
    ch = int(input("Find \n1. Top 3 Performers \n2. Bottom 3 Performers\n"))
    if ch == 1:
        print("\n---------  Top 3 Performers are  ----------\n")
        sorted_list = find_top_performer(sheets)
        save_data(sorted_list)
    elif ch == 2:
        print("\n---------  Bottom 3 Performers are  ----------\n")
        sorted_list = find_bottom_performer(sheets)
        save_data(sorted_list)
    else:
        print("Wrong Choice\n")
    for emp in sorted_list:
        print_obj(emp)
